import openpyxl
import numpy as np
import math
from PIL import Image
from numpy import asarray
from scipy.stats import skew, kurtosis
import skimage.feature
import xlsxwriter


# Create Red File
outWorkbookR = xlsxwriter.Workbook("RedChannel.xlsx")
outSheetR = outWorkbookR.add_worksheet()

# Write Headers
outSheetR.write("A1", "MeanR")
outSheetR.write("B1", "VarianceR")
outSheetR.write("C1", "SkewnessR")
outSheetR.write("D1", "KurtosisR")
outSheetR.write("E1", "ContrastR")
outSheetR.write("F1", "EntropyR")
outSheetR.write("G1", "EnergyR")
outSheetR.write("H1", "HomoR")
outSheetR.write("I1", "CorreR")
outWorkbookR.close()


# Create Green File
outWorkbookG = xlsxwriter.Workbook("GreenChannel.xlsx")
outSheetG = outWorkbookG.add_worksheet()

# Write Headers for Green File
outSheetG.write("A1", "MeanG")
outSheetG.write("B1", "VarianceG")
outSheetG.write("C1", "SkewnessG")
outSheetG.write("D1", "KurtosisG")
outSheetG.write("E1", "ContrastG")
outSheetG.write("F1", "EntropyG")
outSheetG.write("G1", "EnergyG")
outSheetG.write("H1", "HomoG")
outSheetG.write("I1", "CorreG")
outWorkbookG.close()

# Create Blue File
outWorkbookB = xlsxwriter.Workbook("BlueChannel.xlsx")
outSheetB = outWorkbookB.add_worksheet()

# Write Headers for Blue file
outSheetB.write("A1", "MeanB")
outSheetB.write("B1", "VarianceB")
outSheetB.write("C1", "SkewnessB")
outSheetB.write("D1", "KurtosisB")
outSheetB.write("E1", "ContrastB")
outSheetB.write("F1", "EntropyB")
outSheetB.write("G1", "EnergyB")
outSheetB.write("H1", "HomoB")
outSheetB.write("I1", "CorreB")
outWorkbookB.close()


import glob

Image_list = []
for filename in glob.glob('/home/jawad/Downloads/FYP/Codes/Code and Data/MyFolder/*.png'):
    im = Image.open(filename)
    Image_list.append(im)

print(len(Image_list))

for i in range(0, len(Image_list)):
    Image = Image_list[i]
    image = asarray(Image)
    R = image[:, :, 0]
    G = image[:, :, 1]
    B = image[:, :, 2]


    # Feature Selection from Red Channel
    print("Calucation nine Statistical features for Red Channel")
    MeanR = np.mean(R);
    print(MeanR)
    VarianceR = np.var(R)
    VarianceR = math.sqrt(VarianceR)
    print(VarianceR)
    SkewnessR = skew(R.reshape(-1))
    print(SkewnessR)
    KurtosisR = kurtosis(R.reshape(-1))
    print(KurtosisR)
    entropyR = skimage.measure.shannon_entropy(R)
    print(entropyR)
    R = skimage.img_as_ubyte(R)
    g1 = skimage.feature.greycomatrix(R, [1], [0], levels=256, symmetric=False, normed=True)
    ContR = skimage.feature.greycoprops(g1, 'contrast')[0][0]
    print(ContR)
    EnergR = skimage.feature.greycoprops(g1, 'energy')[0][0]
    print(EnergR)
    HomoR = skimage.feature.greycoprops(g1, 'homogeneity')[0][0]
    print(HomoR)
    CorreR = skimage.feature.greycoprops(g1, 'correlation')[0][0]
    print(CorreR)

    # Feature Selection from Green Channel
    print("Calculate Statistical Features for Green Channel")
    MeanG = np.mean(G);
    print(MeanG)
    VarianceG = np.var(G)
    VarianceG = math.sqrt(VarianceG)
    print(VarianceG)
    SkewnessG = skew(G.reshape(-1))
    print(SkewnessG)
    KurtosisG = kurtosis(G.reshape(-1))
    print(KurtosisG)
    entropyG = skimage.measure.shannon_entropy(G)
    print(entropyG)
    G = skimage.img_as_ubyte(G)
    g2 = skimage.feature.greycomatrix(G, [1], [0], levels=256, symmetric=False, normed=True)
    ContG = skimage.feature.greycoprops(g2, 'contrast')[0][0]
    print(ContG)
    EnergG = skimage.feature.greycoprops(g2, 'energy')[0][0]
    print(EnergG)
    HomoG = skimage.feature.greycoprops(g2, 'homogeneity')[0][0]
    print(HomoG)
    CorreG = skimage.feature.greycoprops(g2, 'correlation')[0][0]
    print(CorreG)

    #  Feature Selection from Blue Channel
    print("Calucation  Statistical Features for Blue Channel")
    MeanB = np.mean(B);
    print(MeanB)
    VarianceB = np.var(B)
    VarianceB = math.sqrt(VarianceB)
    print(VarianceB)
    SkewnessB= skew(B.reshape(-1))
    print(SkewnessB)
    KurtosisB = kurtosis(B.reshape(-1))
    print(KurtosisB)
    entropyB = skimage.measure.shannon_entropy(B)
    print(entropyB)
    B = skimage.img_as_ubyte(B)
    g3 = skimage.feature.greycomatrix(B, [1], [0], levels=256, symmetric=False, normed=True)
    ContB = skimage.feature.greycoprops(g3, 'contrast')[0][0]
    print(ContB)
    EnergB = skimage.feature.greycoprops(g3, 'energy')[0][0]
    print(EnergB)
    HomoB = skimage.feature.greycoprops(g3, 'homogeneity')[0][0]
    print(HomoB)
    CorreB = skimage.feature.greycoprops(g3, 'correlation')[0][0]
    print(CorreB)

    print("How many Times")
    print(i)


    # Declare Data
    values_R = [MeanR, VarianceR, SkewnessR, KurtosisR, entropyR, ContR, EnergR, HomoR, CorreR]
    values_G = [MeanG, VarianceG, SkewnessG, KurtosisG, entropyG, ContG, EnergG, HomoG, CorreG]
    values_B = [MeanB, VarianceB, SkewnessB, KurtosisB, entropyB, ContB, EnergB, HomoB, CorreB]

    # Write Red data to file
    outWorkbookR = openpyxl.load_workbook("RedChannel.xlsx")
    outSheetR = outWorkbookR.active
    outSheetR.append(values_R)
    outWorkbookR.save(filename="RedChannel.xlsx")

    # Write Green data to file
    outWorkbookG = openpyxl.load_workbook("GreenChannel.xlsx")
    outSheetG = outWorkbookG.active
    outSheetG.append(values_G)
    outWorkbookG.save(filename="GreenChannel.xlsx")

    # Write Blue data to file
    outWorkbookB = openpyxl.load_workbook("BlueChannel.xlsx")
    outSheetB = outWorkbookB.active
    outSheetB.append(values_B)
    outWorkbookB.save(filename="BlueChannel.xlsx")


